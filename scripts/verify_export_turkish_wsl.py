#!/usr/bin/env python3
"""Verify exported XLSX headers are valid Turkish (no mojibake)."""

from __future__ import annotations

import os
import sys
from io import BytesIO
from pathlib import Path


def _init_frappe() -> None:
    import frappe

    site = os.environ.get("FRAPPE_SITE", "at.localhost")
    sites_path = os.environ.get("FRAPPE_SITES_PATH")
    if not sites_path:
        cwd = Path.cwd()
        if (cwd / site).exists():
            sites_path = str(cwd)
        else:
            bench_root = os.environ.get("FRAPPE_BENCH_ROOT", "")
            if bench_root and (Path(bench_root) / "sites" / site).exists():
                sites_path = str(Path(bench_root) / "sites")
    if sites_path:
        frappe.init(site=site, sites_path=sites_path)
    else:
        frappe.init(site=site)


def main() -> int:
    import frappe
    from openpyxl import load_workbook

    from acentem_takipte.acentem_takipte.api.list_exports import download_export

    _init_frappe()
    frappe.connect()
    frappe.set_user("Administrator")
    frappe.local.lang = "tr"
    frappe.local.response = frappe._dict()

    download_export(screen="customer_list", export_format="xlsx", limit=5, filename="tr_test")
    response = frappe.local.response
    workbook = load_workbook(BytesIO(response.get("filecontent") or b""))
    sheet = workbook.active
    headers = [sheet.cell(row=5, column=index).value for index in range(1, 10)]
    markers = ("Ã", "Ä", "Å")
    bad = [header for header in headers if header and any(marker in str(header) for marker in markers)]

    print("headers:", headers)
    if bad:
        print("MOJIBAKE FOUND:", bad)
        return 1
    if "Müşteri" not in headers and "MÃ¼ÅŸteri" not in str(headers):
        print("Expected Turkish customer header in export")
        return 1
    print("TURKISH EXPORT OK")
    return 0


if __name__ == "__main__":
    sys.exit(main())
