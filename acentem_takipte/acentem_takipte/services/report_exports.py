from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any

import frappe
from frappe.utils.pdf import get_pdf
from openpyxl import Workbook
from openpyxl.styles import Font

from acentem_takipte.acentem_takipte.services.export_payload_utils import coerce_columns, coerce_filters, coerce_rows, normalize_title

REPORT_TITLES = {
    "policy_list": {"tr": "Poliçe Listesi Raporu", "en": "Policy List Report"},
    "payment_status": {"tr": "Tahsilat Durumu Raporu", "en": "Payment Status Report"},
    "renewal_performance": {"tr": "Yenileme Performans Raporu", "en": "Renewal Performance Report"},
    "claim_loss_ratio": {"tr": "Hasar Prim Oranı Raporu", "en": "Claim Loss Ratio Report"},
    "agent_performance": {"tr": "Acente Üretim Karnesi", "en": "Agency Performance Scorecard"},
    "customer_segmentation": {"tr": "Müşteri Segmentasyon Raporu", "en": "Customer Segmentation Report"},
    "communication_operations": {"tr": "İletişim Operasyonları Raporu", "en": "Communication Operations Report"},
    "reconciliation_operations": {"tr": "Mutabakat Operasyonları Raporu", "en": "Reconciliation Operations Report"},
    "claims_operations": {"tr": "Hasar Operasyonları Raporu", "en": "Claims Operations Report"},
}


def build_export_filename(export_key: str, export_format: str) -> str:
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    raw_key = str(export_key or "").strip() or "report"
    safe_key = "".join(char if char.isalnum() or char in {"_", "-"} else "_" for char in raw_key) or "report"
    extension = "pdf" if str(export_format or "").strip().lower() == "pdf" else "xlsx"
    return f"{safe_key}_{timestamp}.{extension}"


def build_report_filename(report_key: str, export_format: str) -> str:
    return build_export_filename(report_key, export_format)


def build_report_title(report_key: str, locale: str = "tr") -> str:
    normalized_key = str(report_key or "").strip()
    entry = REPORT_TITLES.get(normalized_key, {})
    normalized_locale = str(locale or "tr").strip() or "tr"
    base_locale = normalized_locale.split("-")[0]
    return entry.get(normalized_locale) or entry.get(base_locale) or entry.get("en") or normalized_key or "Report"


def render_tabular_pdf(
    *,
    title: str,
    columns: list[str],
    rows: list[dict[str, Any]],
    filters: dict[str, Any],
) -> bytes:
    safe_title = _normalize_title(title)
    safe_columns = _coerce_columns(columns)
    safe_rows = _coerce_rows(rows)
    safe_filters = _coerce_filters(filters)
    html = frappe.render_template(
        """
        <html>
          <head>
            <style>
              body { font-family: Arial, sans-serif; font-size: 11px; color: #0f172a; }
              .header { margin-bottom: 16px; }
              .title { font-size: 18px; font-weight: 700; }
              .meta { margin-top: 6px; color: #475569; font-size: 10px; }
              table { width: 100%; border-collapse: collapse; margin-top: 16px; }
              th, td { border: 1px solid #cbd5e1; padding: 6px 8px; text-align: left; vertical-align: top; }
              th { background: #e2e8f0; font-weight: 700; }
              .summary { margin-top: 10px; font-size: 10px; color: #334155; }
            </style>
          </head>
          <body>
            <div class="header">
              <div class="title">{{ report_title }}</div>
              <div class="meta">Filters: {{ filters }}</div>
              <div class="summary">Total rows: {{ rows|length }}</div>
            </div>
            <table>
              <thead>
                <tr>
                  {% for column in columns %}
                  <th>{{ column }}</th>
                  {% endfor %}
                </tr>
              </thead>
              <tbody>
                {% for row in rows %}
                <tr>
                  {% for column in columns %}
                  <td>{{ row.get(column, "") }}</td>
                  {% endfor %}
                </tr>
                {% endfor %}
              </tbody>
            </table>
          </body>
        </html>
        """,
        {
            "report_title": safe_title,
            "columns": safe_columns,
            "rows": safe_rows,
            "filters": safe_filters,
        },
    )
    return get_pdf(html)


def render_report_pdf(
    *,
    report_key: str,
    columns: list[str],
    rows: list[dict[str, Any]],
    filters: dict[str, Any],
    locale: str = "tr",
) -> bytes:
    return render_tabular_pdf(
        title=build_report_title(report_key, locale),
        columns=columns,
        rows=rows,
        filters=filters,
    )


def render_tabular_xlsx(
    *,
    title: str,
    columns: list[str],
    rows: list[dict[str, Any]],
    filters: dict[str, Any],
) -> bytes:
    safe_title = _normalize_title(title)
    safe_columns = _coerce_columns(columns)
    safe_rows = _coerce_rows(rows)
    safe_filters = _coerce_filters(filters)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Report"

    sheet["A1"] = safe_title
    sheet["A1"].font = Font(bold=True, size=14)
    sheet["A2"] = f"Filters: {safe_filters}"
    sheet["A3"] = f"Total rows: {len(safe_rows)}"

    header_row = 5
    for index, column in enumerate(safe_columns, start=1):
        cell = sheet.cell(row=header_row, column=index, value=column)
        cell.font = Font(bold=True)

    for row_index, row in enumerate(safe_rows, start=header_row + 1):
        for column_index, column in enumerate(safe_columns, start=1):
            sheet.cell(row=row_index, column=column_index, value=row.get(column))

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def render_report_xlsx(
    *,
    report_key: str,
    columns: list[str],
    rows: list[dict[str, Any]],
    filters: dict[str, Any],
    locale: str = "tr",
) -> bytes:
    return render_tabular_xlsx(
        title=build_report_title(report_key, locale),
        columns=columns,
        rows=rows,
        filters=filters,
    )


def _normalize_title(title: Any) -> str:
    return normalize_title(title, "Report")


def _coerce_columns(columns: Any) -> list[str]:
    return coerce_columns(columns)


def _coerce_rows(rows: Any) -> list[dict[str, Any]]:
    return coerce_rows(rows)


def _coerce_filters(filters: Any) -> dict[str, Any]:
    return coerce_filters(filters)

