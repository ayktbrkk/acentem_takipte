from __future__ import annotations

from io import BytesIO
from typing import Any

from acentem_takipte.acentem_takipte.services.data_import.parsers.base import ParsedSheet


def list_xlsx_sheet_names(data: bytes) -> list[str]:
    from openpyxl import load_workbook

    workbook = load_workbook(filename=BytesIO(data), read_only=True, data_only=True)
    return list(workbook.sheetnames)


def parse_xlsx_bytes(
    *,
    data: bytes,
    sheet_name: str | None = None,
    limit: int = 200,
) -> ParsedSheet:
    from openpyxl import load_workbook

    workbook = load_workbook(filename=BytesIO(data), read_only=True, data_only=True)
    sheet_names = list(workbook.sheetnames)
    if not sheet_names:
        return ParsedSheet()

    target_sheet = str(sheet_name or "").strip() or sheet_names[0]
    if target_sheet not in sheet_names:
        target_sheet = sheet_names[0]

    worksheet = workbook[target_sheet]
    row_iter = worksheet.iter_rows(values_only=True)
    header_cells = next(row_iter, None) or []
    headers = [_cell_text(cell) for cell in header_cells if _cell_text(cell)]

    rows: list[dict[str, str]] = []
    safe_limit = max(int(limit), 1)
    for index, cells in enumerate(row_iter):
        if index >= safe_limit:
            break
        if not any(cell not in (None, "") for cell in cells):
            continue
        row = {}
        for col_index, header in enumerate(headers):
            row[header] = _cell_text(cells[col_index] if col_index < len(cells) else "")
        rows.append(row)

    workbook.close()
    return ParsedSheet(headers=headers, rows=rows, sheet_names=sheet_names, active_sheet=target_sheet)


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        try:
            return str(value.date()) if hasattr(value, "date") else value.isoformat()
        except Exception:
            pass
    return str(value).strip()
