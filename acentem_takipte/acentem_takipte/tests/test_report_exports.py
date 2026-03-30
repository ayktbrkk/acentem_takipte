from io import BytesIO
from types import SimpleNamespace

from openpyxl import load_workbook

from acentem_takipte.acentem_takipte.services import report_exports
from acentem_takipte.acentem_takipte.services.report_exports import build_export_filename, build_report_title
from acentem_takipte.acentem_takipte.utils.i18n import translate_text


def test_build_report_title_includes_extended_report_catalog():
    assert build_report_title("communication_operations", "tr") == translate_text("Communication Operations Report", "tr")
    assert build_report_title("reconciliation_operations", "en") == "Reconciliation Operations Report"
    assert build_report_title("claims_operations", "tr") == translate_text("Claims Operations Report", "tr")


def test_build_report_title_uses_full_locale_then_base_locale():
    assert build_report_title("policy_list", "tr-TR") == translate_text("Policy List Report", "tr-TR")


def test_build_report_title_defaults_to_english():
    assert build_report_title("policy_list") == "Policy List Report"


def test_build_report_title_trims_report_key():
    assert build_report_title(" policy_list ", "tr") == translate_text("Policy List Report", "tr")


def test_build_export_filename_uses_report_fallback_for_blank_key():
    filename = build_export_filename("   ", "xlsx")

    assert filename.startswith("report_")


def test_build_export_filename_trims_export_format():
    filename = build_export_filename("policy_list", " PDF ")

    assert filename.endswith(".pdf")


def test_build_report_title_trims_locale():
    assert build_report_title("policy_list", " tr-TR ") == translate_text("Policy List Report", "tr-TR")


def test_render_tabular_pdf_coerces_invalid_shapes(monkeypatch):
    monkeypatch.setattr(report_exports.frappe, "local", SimpleNamespace(lang="tr"), raising=False)
    captured = {}
    monkeypatch.setattr(
        report_exports.frappe,
        "render_template",
        lambda template, context: captured.update(context) or "<html></html>",
    )
    monkeypatch.setattr(report_exports, "get_pdf", lambda html: b"pdf")

    result = report_exports.render_tabular_pdf(
        title="   ",
        columns="name",
        rows={"name": "POL-001"},
        filters="status=Open",
    )

    assert result == b"pdf"
    assert captured["report_title"] == translate_text("Report", "tr")
    assert captured["columns"] == ["name"]
    assert captured["rows"] == []
    assert captured["filters"] == {}
    assert captured["filters_label"] == translate_text("Filters:", "tr")
    assert captured["total_rows_label"] == translate_text("Total rows:", "tr")


def test_render_tabular_xlsx_coerces_invalid_shapes():
    payload = report_exports.render_tabular_xlsx(
        title="   ",
        columns="name",
        rows={"name": "POL-001"},
        filters="status=Open",
    )

    workbook = load_workbook(BytesIO(payload))
    sheet = workbook.active

    assert sheet["A1"].value == "Report"
    assert sheet["A2"].value == "Filters: {}"
    assert sheet["A3"].value == "Total rows: 0"


def test_render_tabular_xlsx_keeps_only_valid_columns_and_rows(monkeypatch):
    monkeypatch.setattr(report_exports.frappe, "local", SimpleNamespace(lang="tr"), raising=False)
    payload = report_exports.render_tabular_xlsx(
        title="Policy List",
        columns=["name", "", "status"],
        rows=[{"name": "POL-001", "status": "Active"}, "bad-row"],
        filters={"status": "Active"},
    )

    workbook = load_workbook(BytesIO(payload))
    sheet = workbook.active

    assert sheet["A1"].value == translate_text("Policy List", "tr")
    assert sheet["A5"].value == "name"
    assert sheet["B5"].value == "status"
    assert sheet["A6"].value == "POL-001"
    assert sheet["B6"].value == "Active"


def test_render_tabular_pdf_accepts_json_filter_string(monkeypatch):
    captured = {}
    monkeypatch.setattr(
        report_exports.frappe,
        "render_template",
        lambda template, context: captured.update(context) or "<html></html>",
    )
    monkeypatch.setattr(report_exports, "get_pdf", lambda html: b"pdf")

    result = report_exports.render_tabular_pdf(
        title="Sample",
        columns=["name"],
        rows=[{"name": "A"}],
        filters='{"status":"Open"}',
    )

    assert result == b"pdf"
    assert captured["filters"] == {"status": "Open"}


def test_render_tabular_xlsx_accepts_string_columns():
    payload = report_exports.render_tabular_xlsx(
        title="Sample",
        columns="name, amount, name",
        rows=[{"name": "A", "amount": 3}],
        filters={},
    )

    workbook = load_workbook(BytesIO(payload))
    sheet = workbook.active

    assert sheet["A5"].value == "name"
    assert sheet["B5"].value == "amount"
    assert sheet.max_column == 2


def test_render_tabular_xlsx_accepts_mapping_rows():
    class MappingRow:
        def items(self):
            return {"name": "A", "amount": 3}.items()

    payload = report_exports.render_tabular_xlsx(
        title="Sample",
        columns=["name", "amount"],
        rows=[MappingRow()],
        filters={},
    )

    workbook = load_workbook(BytesIO(payload))
    sheet = workbook.active

    assert sheet["A6"].value == "A"
    assert sheet["B6"].value == 3

